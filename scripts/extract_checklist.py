#!/usr/bin/env python
"""Trích checklist thẩm định sizing từ file Excel ra Markdown.

Checklist là tài liệu sống — đơn vị thẩm định sẽ phát hành bản mới. Script này
để lần sau chạy lại thay vì chép tay, và để thấy ngay bản mới thêm/bớt mục nào.

Cách dùng
---------
    uv run python scripts/extract_checklist.py "Checklist sizing cap phat tai nguyen.xlsx"
    uv run python scripts/extract_checklist.py <file.xlsx> -o docs/rules/.tmp-checklist/items.md

Cấu trúc file nguồn (đã khảo sát bản 2026-08-25)
------------------------------------------------
    A: TT          B: Hạng mục
    C: Tham chiếu theo tài liệu sizing (Trang, mục bao nhiêu)   <- đơn vị yêu cầu điền
    D: Đánh giá tài liệu kèm SR   E: Kết quả   F: OK/NOK        <- đơn vị thẩm định điền
    G: Ghi chú     H: Tài liệu tham chiếu

Cột G thường chứa **tiêu chí đạt viết bằng lời của chính người thẩm định** — đó là
phần quý nhất cho mục 0.4, nên script giữ nguyên văn, không rút gọn.
"""

from __future__ import annotations

import argparse
import pathlib
import re
import sys

try:
    import openpyxl
except ImportError:  # pragma: no cover
    sys.exit("Thiếu openpyxl. Cài bằng: uv add openpyxl  (hoặc pip install openpyxl)")

LEVEL1 = re.compile(r"^[A-Z]+$")          # A, I, II, III
LEVEL2 = re.compile(r"^\d+\.\d+$")        # 1.1, 2.10
LEVEL3 = re.compile(r"^\d+\.\d+\.\d+$")   # 3.1.1


def clean(v) -> str:
    if v is None:
        return ""
    return re.sub(r"\s*\n\s*", " / ", str(v)).strip()


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__.split("\n")[0])
    ap.add_argument("xlsx", type=pathlib.Path)
    ap.add_argument("-o", "--out", type=pathlib.Path)
    args = ap.parse_args()

    if not args.xlsx.is_file():
        sys.exit(f"Không tìm thấy file: {args.xlsx}")

    ws = openpyxl.load_workbook(args.xlsx, data_only=True).worksheets[0]

    rows, orphans = [], []
    counts = {1: 0, 2: 0, 3: 0}

    for r in ws.iter_rows(min_row=3, max_row=ws.max_row, max_col=8):
        tt, muc = clean(r[0].value), clean(r[1].value)
        ghi_chu, tham_chieu = clean(r[6].value), clean(r[7].value)
        if not tt and not muc:
            continue

        if not tt:  # dòng nội dung bị sót số thứ tự
            orphans.append((r[1].row, muc))
            level = 0
        elif LEVEL3.match(tt):
            level = 3
        elif LEVEL2.match(tt):
            level = 2
        elif LEVEL1.match(tt):
            level = 1
        else:
            level = 2
        if level:
            counts[level] += 1
        rows.append((r[1].row, level, tt, muc, ghi_chu, tham_chieu))

    lines = [
        "| Dòng | TT | Hạng mục | Ghi chú (tiêu chí đạt, nguyên văn) | Tài liệu tham chiếu |",
        "|---|---|---|---|---|",
    ]
    for excel_row, level, tt, muc, ghi_chu, tham_chieu in rows:
        name = f"**{muc}**" if level in (1, 2) and not LEVEL3.match(tt or "") else muc
        lines.append(
            f"| {excel_row} | {tt or '_(sót)_'} | {name} | {ghi_chu} | {tham_chieu} |"
        )

    body = "\n".join(lines)
    if args.out:
        args.out.parent.mkdir(parents=True, exist_ok=True)
        args.out.write_text(body + "\n", encoding="utf-8")
        print(f"-> {args.out}")

    print(f"{args.xlsx.name}: {ws.max_row} dòng")
    print(f"  cấp 1 (A/I/II/III): {counts[1]}")
    print(f"  cấp 2 (x.y)       : {counts[2]}")
    print(f"  cấp 3 (x.y.z)     : {counts[3]}")
    if orphans:
        print(f"  CẢNH BÁO: {len(orphans)} dòng có nội dung nhưng KHÔNG có số TT")
        for excel_row, muc in orphans:
            print(f"    dòng {excel_row}: {muc[:70]}")
    if not args.out:
        print()
        print(body)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
